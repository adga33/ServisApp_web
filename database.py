import os
import time
import shutil
import datetime
import pandas as pd
import streamlit as st
from config import EXCEL_FILE, SHEETS, COLUMNS
from openpyxl import load_workbook

EXCEL_FILE = "servisi.xlsx"

def save_sheet(sheet_name, df):
    """Zamjenjuje postojeći sheet novim DataFrameom."""
    book = load_workbook(EXCEL_FILE)

    # Ako sheet postoji – obriši ga
    if sheet_name in book.sheetnames:
        std = book[sheet_name]
        book.remove(std)

    # Kreiraj novi sheet
    book.create_sheet(sheet_name)
    book.save(EXCEL_FILE)

    # Upis DataFramea
    with pd.ExcelWriter(EXCEL_FILE, engine="openpyxl", mode="a", if_sheet_exists="replace") as writer:
        df.to_excel(writer, sheet_name=sheet_name, index=False)

# -----------------------------
#  Ensure Excel exists
# -----------------------------
def ensure_excel_exists():
    if not os.path.exists(EXCEL_FILE):
        writer = pd.ExcelWriter(EXCEL_FILE, engine="openpyxl")
        for sheet in SHEETS:
            df = pd.DataFrame(columns=COLUMNS)
            df.to_excel(writer, sheet_name=sheet, index=False)
        writer.close()


# -----------------------------
#  Load sheet
# -----------------------------
def load_sheet(sheet):
    ensure_excel_exists()
    try:
        df = pd.read_excel(EXCEL_FILE, sheet_name=sheet)
        return df
    except Exception:
        st.error("Nešto je pošlo krivo pri učitavanju podataka.")
        return pd.DataFrame(columns=COLUMNS)


# -----------------------------
#  Append row
# -----------------------------
def append_row(sheet, data):
    df = load_sheet(sheet)
    df.loc[len(df)] = data

    try:
        with pd.ExcelWriter(EXCEL_FILE, engine="openpyxl", mode="a", if_sheet_exists="replace") as writer:
            df.to_excel(writer, sheet_name=sheet, index=False)
    except Exception:
        st.error("Nešto je pošlo krivo pri spremanju podataka.")


# -----------------------------
#  Backup Excel
# -----------------------------
def backup_excel():
    try:
        timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
        backup_name = f"backup_excel_baza_{timestamp}.xlsx"
        shutil.copy(EXCEL_FILE, backup_name)
        st.success("Backup napravljen.")
    except Exception:
        st.error("Nešto je pošlo krivo pri izradi backup-a.")


# -----------------------------
#  Cleanup old backups
# -----------------------------
def cleanup_old_backups(days=30):
    now = time.time()
    for file in os.listdir():
        if file.startswith("backup_excel_baza_") and file.endswith(".xlsx"):
            if os.stat(file).st_mtime < now - days * 86400:
                os.remove(file)

